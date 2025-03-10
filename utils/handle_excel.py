import openpyxl

from pathlib import Path

# 获取项目根目录
BASE_DIR = Path(__file__).resolve().parent.parent


# 读取excel
class ReadExcel:
    def __init__(self, file_name):
        self.workbook = openpyxl.load_workbook(file_name)
        self.worksheet = self.workbook.active

    def get_data(self):
        data = []
        for row in self.worksheet.iter_rows():
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            data.append(row_value)
        return data


class WriteExcel:
    def __init__(self,file_path,data):
        self.file_path = file_path
        self.data = data
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

    def write_to_excel(self):
        for row in self.data:
            self.worksheet.append(row)
        self.workbook.save(self.file_path)

if __name__ == "__main__":
    read_excel_obj = ReadExcel(str(BASE_DIR / "students.xlsx"))
    result = read_excel_obj.get_data()
    print(result)
    # write_excel_obj = WriteExcel()
    # write_excel_obj.write_to_excel()